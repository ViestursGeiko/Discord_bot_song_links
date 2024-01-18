import discord
from discord.ext import commands
from scrape import discord_url
import openpyxl

intents = discord.Intents.default()
intents.messages = True
intents.message_content = True

TOKEN = "MTE5NzEyMjg3NDA5NjUwNDkyMg.G_JHdB.v9j_LNX1mTmVs0lMvDpCHqrHEqgtXEZJe2Q180"
#dod prefix komandam
bot = commands.Bot(command_prefix='!', intents=intents)

#piesledz un izvada pazinojuma par to ka bots ir online
@bot.event
async def on_ready():
   print(f'Logged in as {bot.user.name}')

#izveidojam komandu uz kuru bots izsauks, scrape skriptu
@bot.command(name='song')
async def song(ctx, *, song_name):
    url = discord_url(song_name)
    await ctx.send(f"Requested song: {url[1]}")
    
    #iegūto linku un nosaukumu no scrape skript ievadam excel failā priekš velākas pārskatīšanas
    wb = openpyxl.load_workbook('song_list.xlsx')
    ws = wb.active

    max_row = ws.max_row
    next_row = ws.max_row +1
    ws.cell(row=next_row, column=2).value=url[1]
    ws.cell(row=next_row, column=1).value=url[0]
    
    wb.save('song_list.xlsx')
    wb.close()

bot.run(TOKEN)
